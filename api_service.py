import re
import shutil
import tempfile
import urllib.request
import uuid
from io import BytesIO
from pathlib import Path

import openpyxl
from fastapi import FastAPI, File, HTTPException, Query, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage

app = FastAPI(title="Excel Image Link Processor", version="1.0.0")

# Allow browser calls from local preview and deployed frontends.
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=False,
    allow_methods=["GET", "POST", "OPTIONS"],
    allow_headers=["*"],
    expose_headers=["X-Changed-Cells", "X-Inserted-Images", "X-Failed-Images", "Content-Disposition"],
)

IMAGE_EXT = re.compile(r"\.(jpg|jpeg|png|gif|webp|bmp|svg)(\?|$)", re.IGNORECASE)
URL_RE = re.compile(r"https?://\S+")


def _normalize_header(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    # Remove separators to make fuzzy matching more tolerant.
    return re.sub(r"[\s_\-()（）【】\[\]：:]+", "", text)


def _find_image_source_column(ws) -> int:
    """
    Find the product-image URL column by header keywords.
    Fallback to column E for compatibility with legacy templates.
    """
    header_keywords = [
        "产品图片",
        "商品图片",
        "图片链接",
        "图片url",
        "imageurl",
        "productimage",
        "mainimage",
        "imagem",
        "imagemproduto",
    ]
    for col in range(1, ws.max_column + 1):
        header = _normalize_header(ws.cell(row=1, column=col).value)
        if not header:
            continue
        if any(k in header for k in header_keywords):
            return col
    return 5


def _find_output_column(ws, source_col: int) -> int:
    """
    Use the nearest empty column to the right of source column.
    If none is empty, append one at the end.
    """
    for col in range(source_col + 1, ws.max_column + 1):
        if ws.cell(row=1, column=col).value in (None, ""):
            return col
    return ws.max_column + 1


def process_workbook(
    input_path: Path,
    output_path: Path,
    max_thumb_size: tuple[int, int] = (220, 220),
    jpeg_quality: int = 60,
    max_rows: int = 0,
    image_timeout: int = 6,
) -> dict:
    wb = openpyxl.load_workbook(input_path)
    ws = wb.active

    changed = 0
    inserted = 0
    failed = 0
    source_col = _find_image_source_column(ws)
    output_col = _find_output_column(ws, source_col)
    output_col_letter = get_column_letter(output_col)
    ws.cell(row=1, column=output_col).value = "首图"
    ws.column_dimensions[output_col_letter].width = 18

    image_dir = output_path.parent / f"images_{uuid.uuid4().hex[:8]}"
    image_dir.mkdir(parents=True, exist_ok=True)

    try:
        end_row = ws.max_row if max_rows <= 0 else min(ws.max_row, max_rows + 1)
        for row in range(2, end_row + 1):
            cell = ws.cell(row=row, column=source_col)
            value = cell.value
            if not isinstance(value, str) or not value.strip():
                continue

            urls = [u.strip(" ,;\"'()[]") for u in URL_RE.findall(value)]
            if not urls:
                continue

            keep = next((u for u in urls if IMAGE_EXT.search(u)), urls[0])
            if value != keep:
                cell.value = keep
                changed += 1

            try:
                req = urllib.request.Request(keep, headers={"User-Agent": "Mozilla/5.0"})
                with urllib.request.urlopen(req, timeout=image_timeout) as resp:
                    image_bytes = resp.read()

                pil_image = PILImage.open(BytesIO(image_bytes))
                if pil_image.mode != "RGB":
                    pil_image = pil_image.convert("RGB")
                pil_image.thumbnail(max_thumb_size, PILImage.Resampling.LANCZOS)

                image_path = image_dir / f"row_{row}.jpg"
                pil_image.save(
                    image_path,
                    format="JPEG",
                    quality=jpeg_quality,
                    optimize=True,
                    progressive=True,
                )

                excel_image = XLImage(str(image_path))
                max_h = 95
                if excel_image.height > max_h:
                    ratio = max_h / excel_image.height
                    excel_image.height = int(excel_image.height * ratio)
                    excel_image.width = int(excel_image.width * ratio)

                ws.add_image(excel_image, f"{output_col_letter}{row}")
                ws.row_dimensions[row].height = 72
                inserted += 1
            except Exception:
                failed += 1

        wb.save(output_path)
    finally:
        shutil.rmtree(image_dir, ignore_errors=True)

    return {"changed": changed, "inserted": inserted, "failed": failed}


@app.get("/health")
def health() -> JSONResponse:
    return JSONResponse({"ok": True})


@app.post("/process")
async def process_excel(
    file: UploadFile = File(...),
    max_rows: int = Query(default=0, ge=0, le=2000),
    image_timeout: int = Query(default=6, ge=2, le=20),
    jpeg_quality: int = Query(default=60, ge=35, le=90),
) -> FileResponse:
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="请上传 .xlsx 文件")

    with tempfile.TemporaryDirectory() as tmp:
        tmp_dir = Path(tmp)
        input_path = tmp_dir / f"in_{uuid.uuid4().hex}.xlsx"
        output_path = tmp_dir / f"out_{uuid.uuid4().hex}.xlsx"

        content = await file.read()
        input_path.write_bytes(content)

        stats = process_workbook(
            input_path,
            output_path,
            jpeg_quality=jpeg_quality,
            max_rows=max_rows,
            image_timeout=image_timeout,
        )
        download_name = file.filename.replace(".xlsx", "_with_images_small.xlsx")

        final_output = Path(tempfile.gettempdir()) / f"{uuid.uuid4().hex}_{download_name}"
        final_output.write_bytes(output_path.read_bytes())

    headers = {
        "X-Changed-Cells": str(stats["changed"]),
        "X-Inserted-Images": str(stats["inserted"]),
        "X-Failed-Images": str(stats["failed"]),
    }
    return FileResponse(
        final_output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=download_name,
        headers=headers,
    )
