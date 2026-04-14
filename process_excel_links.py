import re
import urllib.request
from io import BytesIO
from pathlib import Path

import openpyxl
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage


INPUT_FILE = Path("User_Products_Export_0414143621.xlsx")
OUTPUT_FILE = Path("User_Products_Export_0414143621_with_images_small.xlsx")
IMAGE_DIR = Path("downloaded_images_tmp")
MAX_THUMB_SIZE = (220, 220)
JPEG_QUALITY = 60


def main() -> None:
    wb = openpyxl.load_workbook(INPUT_FILE)
    ws = wb.active

    image_ext = re.compile(r"\.(jpg|jpeg|png|gif|webp|bmp|svg)(\?|$)", re.IGNORECASE)
    changed = 0
    inserted = 0
    failed = 0
    IMAGE_DIR.mkdir(exist_ok=True)
    ws["F1"] = "首图"
    ws.column_dimensions["F"].width = 18

    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=5)  # E column
        value = cell.value
        if not isinstance(value, str) or not value.strip():
            continue

        urls = [u.strip(" ,;\"'()[]") for u in re.findall(r"https?://\S+", value)]
        if not urls:
            continue

        keep = next((u for u in urls if image_ext.search(u)), urls[0])
        if value != keep:
            cell.value = keep
            changed += 1

        try:
            req = urllib.request.Request(
                keep,
                headers={"User-Agent": "Mozilla/5.0"},
            )
            with urllib.request.urlopen(req, timeout=15) as resp:
                image_bytes = resp.read()

            pil_image = PILImage.open(BytesIO(image_bytes))
            if pil_image.mode != "RGB":
                pil_image = pil_image.convert("RGB")
            pil_image.thumbnail(MAX_THUMB_SIZE, PILImage.Resampling.LANCZOS)

            image_path = IMAGE_DIR / f"row_{row}.jpg"
            pil_image.save(
                image_path,
                format="JPEG",
                quality=JPEG_QUALITY,
                optimize=True,
                progressive=True,
            )

            excel_image = XLImage(str(image_path))
            max_h = 95
            if excel_image.height > max_h:
                ratio = max_h / excel_image.height
                excel_image.height = int(excel_image.height * ratio)
                excel_image.width = int(excel_image.width * ratio)

            ws.add_image(excel_image, f"F{row}")
            ws.row_dimensions[row].height = 72
            inserted += 1
        except Exception:
            failed += 1

    wb.save(OUTPUT_FILE)
    print(f"处理完成，更新单元格数量: {changed}")
    print(f"插入图片数量: {inserted}")
    print(f"下载/插入失败数量: {failed}")
    print(f"输出文件: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
